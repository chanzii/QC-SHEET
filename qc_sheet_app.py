import streamlit as st
import os, subprocess, shutil, re, json, base64, requests
from io import BytesIO
from pathlib import Path
from urllib.parse import quote as url_quote       # ← URL 인코딩용
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# -------------------------------------------------------
# 기본 설정
# -------------------------------------------------------
st.set_page_config(page_title="QC시트 자동 생성기", layout="centered")
st.title(" QC시트 생성기 ")

# -------------------------------------------------------
# 경로 설정
# -------------------------------------------------------
BASE_DIR     = "uploaded"
SPEC_DIR     = os.path.join(BASE_DIR, "spec")
TEMPLATE_DIR = os.path.join(BASE_DIR, "template")
IMAGE_DIR    = os.path.join(BASE_DIR, "image")
for folder in (SPEC_DIR, TEMPLATE_DIR, IMAGE_DIR):
    os.makedirs(folder, exist_ok=True)

# -------------------------------------------------------
# GitHub API 설정 (토큰·저장소 정보는 secrets.toml 또는 Cloud Secrets에!)
# -------------------------------------------------------
GH_TOKEN  = st.secrets["GH_TOKEN"]
GH_REPO   = st.secrets["GH_REPO"]            # ex) chanzii/QC-SHEET
GH_BRANCH = st.secrets.get("GH_BRANCH", "main")
GH_API    = f"https://api.github.com/repos/{GH_REPO}/contents"
HEADERS   = {"Authorization": f"token {GH_TOKEN}",
             "Accept": "application/vnd.github+json"}

# -------------------------------------------------------
# GitHub ▶︎ 로컬 동기화 (앱 시작 시 1회)
# -------------------------------------------------------
REPO_LOCAL = Path("repo_cache")   # 임시 클론 위치

def sync_repo():
    """GitHub 저장소에 있는 spec/template/image 폴더를 uploaded/ 로 복원"""
    repo_url = f"https://{GH_TOKEN}@github.com/{GH_REPO}.git"
    try:
        if REPO_LOCAL.exists():
            subprocess.run(["git", "-C", str(REPO_LOCAL), "pull", "--quiet"], check=True)
        else:
            subprocess.run(["git", "clone", "--depth", "1", "--branch", GH_BRANCH,
                            repo_url, str(REPO_LOCAL)], check=True)
    except subprocess.CalledProcessError as e:
        st.warning(f"⚠️ GitHub 동기화 실패: {e}")
        return

    for name in ("spec", "template", "image"):
        src = REPO_LOCAL / name
        dst = Path(BASE_DIR) / name
        if src.exists():
            for f in src.iterdir():
                if f.is_file():
                    shutil.copy2(f, dst / f.name)

sync_repo()   # ★ 앱 부팅 시 1회 실행

# -------------------------------------------------------
# GitHub 업로드 & 삭제 유틸
# -------------------------------------------------------
def github_commit(local_path: str, repo_rel_path: str):
    """local_path → GitHub (생성/덮어쓰기)"""
    with open(local_path, "rb") as f:
        content = base64.b64encode(f.read()).decode()

    # sha 확인
    sha = None
    r = requests.get(f"{GH_API}/{url_quote(repo_rel_path)}",
                     params={"ref": GH_BRANCH}, headers=HEADERS)
    if r.status_code == 200:
        sha = r.json().get("sha")

    payload = {"message": f"upload {repo_rel_path}",
               "content": content,
               "branch": GH_BRANCH}
    if sha:
        payload["sha"] = sha

    r = requests.put(f"{GH_API}/{url_quote(repo_rel_path)}",
                     headers=HEADERS, data=json.dumps(payload))
    if r.status_code in (200, 201):
        st.toast("✅ GitHub 커밋 완료", icon="🎉")
    else:
        st.error(f"❌ GitHub 커밋 실패: {r.status_code} {r.json().get('message')}")

def github_delete(repo_rel_path: str) -> bool:
    """GitHub에서 파일 삭제, 성공 시 True"""
    api = f"{GH_API}/{url_quote(repo_rel_path)}"
    r = requests.get(api, params={"ref": GH_BRANCH}, headers=HEADERS)
    if r.status_code != 200:
        return False
    sha = r.json().get("sha")
    payload = {"message": f"delete {repo_rel_path}",
               "sha": sha,
               "branch": GH_BRANCH}
    r = requests.delete(api, headers=HEADERS, json=payload)
    return r.status_code in (200, 204)

# -------------------------------------------------------
# 업로드 & 삭제 UI
# -------------------------------------------------------
def uploader(label, subfolder, repo_folder, multiple):
    files = st.file_uploader(label,
                             type=["xlsx", "png", "jpg", "jpeg"],
                             accept_multiple_files=multiple)
    if files:
        for f in files:
            local_path = os.path.join(subfolder, f.name)
            with open(local_path, "wb") as fp:
                fp.write(f.getbuffer())
            github_commit(local_path, f"{repo_folder}/{f.name}")
        st.success("✅ 업로드 & GitHub 커밋 완료!")

st.subheader("📁 파일 업로드 및 관리")
col_spec, col_tmp, col_img = st.columns(3)
with col_spec:
    uploader("🧾 스펙 엑셀 업로드", SPEC_DIR, "spec", multiple=True)
with col_tmp:
    uploader("📄 QC시트 양식 업로드", TEMPLATE_DIR, "template", multiple=False)
with col_img:
    uploader("🖼️ 서명/로고 업로드", IMAGE_DIR, "image", multiple=True)

with st.expander("🗑️ 업로드된 파일 삭제하기"):
    mapping = [("스펙", SPEC_DIR, "spec"),
               ("양식", TEMPLATE_DIR, "template"),
               ("이미지", IMAGE_DIR, "image")]
    for label, local_dir, repo_folder in mapping:
        files = os.listdir(local_dir)
        if not files:
            continue

        st.markdown(f"**{label} 파일**")
        for fn in files:
            cols = st.columns([7,1,1])
            cols[0].write(fn)
            # ⬇️ 다운로드
            cols[1].download_button("⬇️",
                                    data=Path(local_dir, fn).read_bytes(),
                                    file_name=fn,
                                    key=f"dl_{local_dir}_{fn}")
            # ❌ 삭제
            if cols[2].button("❌", key=f"del_{local_dir}_{fn}"):
                # 1) 로컬(사용 폴더) 삭제
                Path(local_dir, fn).unlink(missing_ok=True)
                # 2) repo_cache 에도 삭제
                (REPO_LOCAL / repo_folder / fn).unlink(missing_ok=True)
                # 3) GitHub 삭제
                ok = github_delete(f"{repo_folder}/{fn}")
                if ok:
                    st.toast("🗑️ GitHub 삭제 완료", icon="✅")
                else:
                    st.error("⚠️ GitHub 삭제 실패 – 토큰 권한·경로 확인")
                st.rerun()

st.markdown("---")

# -------------------------------------------------------
# QC시트 생성 파트
# -------------------------------------------------------
st.subheader("📄 QC시트 생성")

spec_files    = os.listdir(SPEC_DIR)
selected_spec = st.selectbox("사용할 스펙 엑셀 선택", spec_files) if spec_files else None

if selected_spec:
    spec_path = os.path.join(SPEC_DIR, selected_spec)
    with open(spec_path, "rb") as f:
        st.download_button("⬇️ 선택한 스펙 파일 다운로드",
                           data=f.read(),
                           file_name=selected_spec,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key=f"dl_spec_{selected_spec}")

style_number  = st.text_input("스타일넘버 입력")
raw_size = st.text_input("사이즈 입력 (예: XS, XL, 5XL, FREE, 28, 90 등)")
selected_size = raw_size.strip()
logo_files    = os.listdir(IMAGE_DIR)
selected_logo = st.selectbox("서명/로고 선택", logo_files) if logo_files else None
language_choice = st.selectbox("측정부위 언어", ["English", "Korean"], index=0)

if st.button("🚀 QC시트 생성"):
    if not selected_spec or not style_number or not selected_logo or not selected_size:
        st.error("⚠️ 필수 값을 확인하세요. (스펙/스타일/사이즈/로고)")
        st.stop()
    template_list = os.listdir(TEMPLATE_DIR)
    if not template_list:
        st.error("⚠️ QC시트 양식이 없습니다. 업로드해주세요.")
        st.stop()

    spec_path     = os.path.join(SPEC_DIR, selected_spec)
    template_path = os.path.join(TEMPLATE_DIR, template_list[0])

    wb_spec = load_workbook(spec_path, data_only=True, read_only=True)
    def matches_style(val, style): return val and style.upper() in str(val).upper()

    ws_spec = next((ws for ws in wb_spec.worksheets
                    if matches_style(ws["A1"].value, style_number)), wb_spec.active)
    if ws_spec is wb_spec.active:
        st.warning("❗ A1에서 스타일넘버를 찾지 못해 첫 시트를 사용합니다.")

    wb_tpl, ws_tpl = load_workbook(template_path), None
    ws_tpl = wb_tpl.active
    ws_tpl["B6"], ws_tpl["G6"] = style_number, selected_size
    ws_tpl.add_image(XLImage(os.path.join(IMAGE_DIR, selected_logo)), "F2")

    header   = list(ws_spec.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    size_map = {str(v).strip(): idx for idx, v in enumerate(header) if v}
    if selected_size not in size_map:
        st.error("⚠️ 사이즈 열이 없습니다.")
        st.stop()
    idx = size_map[selected_size]

    rows = list(ws_spec.iter_rows(min_row=3, values_only=True))
    data, i = [], 0
    while i < len(rows):
        part = str(rows[i][1]).strip() if rows[i][1] else ""
        val  = rows[i][idx]
        if language_choice == "English":
            if re.search(r"[A-Za-z]", part) and val is not None:
                data.append((part, val)); i += 1; continue
        else:  # Korean
            if re.search(r"[A-Za-z]", part) and val is not None and i + 1 < len(rows):
                kr = str(rows[i+1][1]).strip() if rows[i+1][1] else ""
                if re.search(r"[가-힣]", kr):
                    data.append((kr, val)); i += 2; continue
            if re.search(r"[가-힣]", part) and val is not None:
                data.append((part, val)); i += 1; continue
        i += 1

    if not data:
        st.error("⚠️ 추출 데이터가 없습니다.")
        st.stop()

    for j, (p, v) in enumerate(data):
        r = 9 + j
        ws_tpl.cell(r, 1, p)
        ws_tpl.cell(r, 2, v)
        ws_tpl.cell(r, 4, f"=IF(C{r}=\"\",\"\",IFERROR(C{r}-B{r},\"\"))")

    out = f"QC_{style_number}_{selected_size}.xlsx"
    buf = BytesIO(); wb_tpl.save(buf); buf.seek(0)

    st.download_button("⬇️ QC시트 다운로드",
                       data=buf.getvalue(),
                       file_name=out,
                       key=f"dl_{out}")
    st.success("✅ QC시트 생성 완료!")
